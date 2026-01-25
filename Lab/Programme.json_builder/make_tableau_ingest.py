#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime as dt
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


IN_PATH = Path("work/normalized.xlsx")
SOURCE_PATH = Path("input/source.xlsx")
OUT_PATH = Path("work/tableau_ingest.xlsx")

DOW_MAP = {
    0: "LU",
    1: "MA",
    2: "ME",
    3: "JE",
    4: "VE",
    5: "SA",
    6: "DI",
}

_ISO_DATE_RE = re.compile(r"^\d{4}([-\/])\d{2}\1\d{2}$")


def _to_date(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        match = _ISO_DATE_RE.match(text)
        if match:
            sep = match.group(1)
            fmt = "%Y-%m-%d" if sep == "-" else "%Y/%m/%d"
            parsed = pd.to_datetime(text, format=fmt, errors="coerce")
        else:
            parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    else:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _format_time(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, dt.datetime):
        hour = value.hour
        minute = value.minute
    elif isinstance(value, dt.time):
        hour = value.hour
        minute = value.minute
    else:
        text = str(value).strip()
        if not text:
            return ""
        compact = text.replace(" ", "")
        if "h" in compact:
            return compact
        parts = compact.split(":")
        if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
            hour = int(parts[0])
            minute = int(parts[1])
        elif compact.isdigit():
            hour = int(compact)
            minute = 0
        else:
            return compact
    return f"{hour}h" if minute == 0 else f"{hour}h{minute:02d}"


def _extract_cm_title(text: str) -> str:
    if not text:
        return ""
    cleaned = str(text).strip()
    for key in ("CM1", "CM2"):
        if cleaned.startswith(key):
            cleaned = cleaned[len(key):].strip()
            break
    if " - " in cleaned:
        cleaned = cleaned.split(" - ", 1)[0].strip()
    return cleaned


def _load_cm_titles() -> dict:
    wb = load_workbook(SOURCE_PATH, data_only=True)
    ws = wb.active
    cm1_raw = ws.cell(row=1, column=5).value or ""
    cm2_raw = ws.cell(row=2, column=5).value or ""
    return {
        "CM1": _extract_cm_title(cm1_raw),
        "CM2": _extract_cm_title(cm2_raw),
    }


def main() -> int:
    df = pd.read_excel(IN_PATH, sheet_name=0, dtype=object).fillna("")
    cm_titles = _load_cm_titles()

    rows = []
    for _, row in df.iterrows():
        date_value = row.get("Date", "")
        date_obj = _to_date(date_value)
        if date_obj:
            dow = DOW_MAP.get(date_obj.weekday(), "")
            day_num = date_obj.day
        else:
            dow = ""
            day_num = ""
        heure = _format_time(row.get("Heure", ""))

        cm_cell = str(row.get("CM", "")).upper()
        cm_keys = []
        for key in ("CM1", "CM2"):
            if key in cm_cell and cm_titles.get(key):
                cm_keys.append(key)
                rows.append([f"{key} - {cm_titles[key]}", "", dow, day_num, heure])

        titre = str(row.get("Titre", "")).strip()
        vovf = str(row.get("VOVF", "")).strip()
        if not vovf:
            vovf = str(row.get("Version", "")).strip()
        if vovf.upper().startswith("VO"):
            titre = f"{titre} - VOST" if titre else titre
        categorie = str(row.get("Categorie", "")).strip()
        if categorie.upper() == "SCOL":
            titre = f"{titre} - SCOL" if titre else titre
        for key in cm_keys:
            titre = f"{titre} + {key}" if titre else titre
        rows.append([titre, "", dow, day_num, heure])

    out_df = pd.DataFrame(rows, columns=["Titre", "", "Jour", "Date", "Heure"])
    out_df.to_excel(OUT_PATH, index=False)

    wb = load_workbook(OUT_PATH)
    ws = wb.active
    align = Alignment(horizontal="center", vertical="center")
    base_font = Font(name="Merriweather", size=13, bold=False)
    bold_font = Font(name="Merriweather", size=13, bold=True)

    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 1:
                cell.font = bold_font
            else:
                cell.font = base_font
            cell.alignment = align

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 40

    wb.save(OUT_PATH)
    print(f"OK: {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
