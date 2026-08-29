#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generate the monthly ingest workbook used for manual copy/paste.

Output columns:
- Titre
- Version (VF, VO ou VOST OCAP)
- Date (ex: "mardi 24 mars")
- Heure (ex: "21h" / "20h30")
- Accessibilité (ex: "AD SME SR G*" / "À vérifier")

Rules:
- Version information is no longer appended to the title; it goes in column B.
- CM rows remain separate and the feature title keeps its "+ CMx" markers.
- School screenings keep their title suffix.
"""

import argparse
import datetime as dt
import json
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

try:
    from .accessibility_lookup import (
        DEFAULT_CACHE_PATH,
        DEFAULT_REPORT_PATH,
        accessibility_for,
        lookup_films,
    )
    from .normalize import extract_cm_catalog
except ImportError:
    from accessibility_lookup import (
        DEFAULT_CACHE_PATH,
        DEFAULT_REPORT_PATH,
        accessibility_for,
        lookup_films,
    )
    from normalize import extract_cm_catalog


BASE_DIR = Path(__file__).resolve().parent
IN_PATH = BASE_DIR / "work/normalized.xlsx"
SOURCE_PATH = BASE_DIR / "input/source.xlsx"
OUT_PATH = BASE_DIR / "work/tableau_ingest.xlsx"

WEEKDAY_FR = {
    0: "lundi",
    1: "mardi",
    2: "mercredi",
    3: "jeudi",
    4: "vendredi",
    5: "samedi",
    6: "dimanche",
}

MONTH_FR = {
    1: "janvier",
    2: "fevrier",
    3: "mars",
    4: "avril",
    5: "mai",
    6: "juin",
    7: "juillet",
    8: "aout",
    9: "septembre",
    10: "octobre",
    11: "novembre",
    12: "decembre",
}

_ISO_DATE_RE = re.compile(r"^\d{4}([-\/])\d{2}\1\d{2}$")


def _to_date(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        match = _ISO_DATE_RE.match(text)
        if match:
            sep = match.group(1)
            fmt = "%Y-%m-%d" if sep == "-" else "%Y/%m/%d"
            parsed = pd.to_datetime(text, format=fmt, errors="coerce")
        else:
            parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    else:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _format_time(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, dt.datetime):
        hour = value.hour
        minute = value.minute
    elif isinstance(value, dt.time):
        hour = value.hour
        minute = value.minute
    else:
        text = str(value).strip()
        if not text:
            return ""
        compact = text.replace(" ", "")
        if "h" in compact:
            return compact
        parts = compact.split(":")
        if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
            hour = int(parts[0])
            minute = int(parts[1])
        elif compact.isdigit():
            hour = int(compact)
            minute = 0
        else:
            return compact
    return f"{hour}h" if minute == 0 else f"{hour}h{minute:02d}"


def _format_full_date(date_obj: dt.date | None) -> str:
    if not date_obj:
        return ""
    weekday = WEEKDAY_FR.get(date_obj.weekday(), "")
    month = MONTH_FR.get(date_obj.month, "")
    if not weekday or not month:
        return ""
    return f"{weekday} {date_obj.day} {month}"


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _format_version(value: object) -> str:
    raw = str(value or "").strip()
    compact = re.sub(r"[^a-z0-9]+", "", _normalize_text(raw))
    if compact == "vostocap":
        return "VOST OCAP"
    if compact.startswith("vo"):
        return "VO"
    if compact.startswith("vf"):
        return "VF"
    return raw


def _is_scolaire(categorie: object) -> bool:
    return bool(re.search(r"\bscol(?:aire)?\b", _normalize_text(categorie)))


def _extract_cm_title(text: str) -> str:
    if not text:
        return ""
    cleaned = str(text).strip()
    match = re.match(r"^(CM\d+)\s*:\s*(.+)$", cleaned)
    if match:
        cleaned = match.group(2).strip()
    else:
        for key in ("CM1", "CM2"):
            if cleaned.startswith(key):
                cleaned = cleaned[len(key):].strip(" :")
                break
    if " - " in cleaned:
        cleaned = cleaned.split(" - ", 1)[0].strip()
    return cleaned


def _load_cm_titles(source_path: Path) -> dict:
    if not source_path.exists():
        return {}
    try:
        raw = pd.read_excel(source_path, sheet_name=0, header=None, dtype=object)
        catalog = extract_cm_catalog(raw)
    except Exception:
        return {}
    return {
        code: str(payload.get("titre", "")).strip()
        for code, payload in catalog.items()
        if str(payload.get("titre", "")).strip()
    }


def _parse_courts_metrages(value: object) -> list[dict]:
    if isinstance(value, list):
        payload = value
    else:
        text = str(value or "").strip()
        if not text:
            return []
        try:
            payload = json.loads(text)
        except (TypeError, ValueError, json.JSONDecodeError):
            return []
    if not isinstance(payload, list):
        return []
    return [item for item in payload if isinstance(item, dict)]


def _cm_titles_from_normalized(df: pd.DataFrame) -> dict[str, str]:
    titles = {}
    for _, row in df.iterrows():
        for court in _parse_courts_metrages(row.get("courts_metrages", "")):
            code = str(court.get("code", "")).strip().upper()
            title = str(court.get("titre", "")).strip()
            if re.fullmatch(r"CM\d+", code) and title:
                titles[code] = title
    return titles


def _cm_refs(value: object) -> list[str]:
    refs = []
    for number in re.findall(r"\bCM\s*(\d+)\b", str(value or ""), flags=re.IGNORECASE):
        code = f"CM{number}"
        if code not in refs:
            refs.append(code)
    return refs


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Génère le tableau mensuel utilisé pour l'ingest."
    )
    parser.add_argument("--input", type=Path, default=IN_PATH)
    parser.add_argument("--source", type=Path, default=SOURCE_PATH)
    parser.add_argument("--output", type=Path, default=OUT_PATH)
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE_PATH)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH)
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Utilise uniquement le dernier cache d'accessibilité.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    df = pd.read_excel(args.input, sheet_name=0, dtype=object).fillna("")
    cm_titles = _cm_titles_from_normalized(df)
    for code, title in _load_cm_titles(args.source).items():
        cm_titles.setdefault(code, title)

    films = []
    for _, row in df.iterrows():
        films.append(
            {
                "title": str(row.get("Titre", "")).strip(),
                "director": str(row.get("Realisateur", "")).strip(),
            }
        )
    for cm_title in cm_titles.values():
        if cm_title:
            films.append({"title": cm_title, "director": ""})

    accessibility, _ = lookup_films(
        films,
        offline=args.offline,
        cache_path=args.cache,
        report_path=args.report,
    )

    rows = []
    for _, row in df.iterrows():
        date_value = row.get("Date", "")
        date_obj = _to_date(date_value)
        date_label = _format_full_date(date_obj)
        heure = _format_time(row.get("Heure", ""))

        cm_keys = _cm_refs(row.get("CM", ""))
        for key in cm_keys:
            if cm_titles.get(key):
                rows.append(
                    [
                        f"{key} - {cm_titles[key]}",
                        "",
                        date_label,
                        heure,
                        accessibility_for(accessibility, cm_titles[key]),
                    ]
                )

        titre = str(row.get("Titre", "")).strip()
        lookup_title = titre
        director = str(row.get("Realisateur", "")).strip()
        vovf = str(row.get("VOVF", "")).strip()
        if not vovf:
            vovf = str(row.get("Version", "")).strip()
        version_label = _format_version(vovf)
        categorie = str(row.get("Categorie", "")).strip()
        if _is_scolaire(categorie):
            titre = f"{titre} - SCOL" if titre else titre
        for key in cm_keys:
            titre = f"{titre} + {key}" if titre else titre
        rows.append(
            [
                titre,
                version_label,
                date_label,
                heure,
                accessibility_for(accessibility, lookup_title, director),
            ]
        )

    out_df = pd.DataFrame(
        rows, columns=["Titre", "Version", "Date", "Heure", "Accessibilité"]
    )

    try:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        out_df.to_excel(args.output, index=False)

        wb = load_workbook(args.output)
        ws = wb.active
        align = Alignment(horizontal="center", vertical="center")
        base_font = Font(name="Merriweather", size=13, bold=False)
        bold_font = Font(name="Merriweather", size=13, bold=True)

        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1:
                    cell.font = bold_font
                else:
                    cell.font = base_font
                cell.alignment = align

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 40

        ws.column_dimensions["A"].width = 56
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 24

        wb.save(args.output)
    except PermissionError as exc:
        raise SystemExit(
            f"Impossible d'ecrire {args.output}. "
            "Ferme le fichier Excel s'il est ouvert, puis relance le script."
        ) from exc

    print(f"OK: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
