#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generate the monthly ingest workbook used for manual copy/paste.

Output columns:
- Titre
- VO
- Date (ex: "mardi 24 mars")
- Heure (ex: "21h" / "20h30")

Rules:
- VO/VOST information is no longer appended to the title; it goes in column B.
- CM rows remain separate and the feature title keeps its "+ CMx" markers.
- School screenings keep their title suffix.
"""

import datetime as dt
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

BASE_DIR = Path(__file__).resolve().parent
IN_PATH = BASE_DIR / "work/normalized.xlsx"
SOURCE_PATH = BASE_DIR / "input/source.xlsx"
OUT_PATH = BASE_DIR / "work/tableau_service.xlsx"

WEEKDAY_FR = {
    0: "lundi",
    1: "mardi",
    2: "mercredi",
    3: "jeudi",
    4: "vendredi",
    5: "samedi",
    6: "dimanche",
}

MONTH_FR = {
    1: "janvier",
    2: "fevrier",
    3: "mars",
    4: "avril",
    5: "mai",
    6: "juin",
    7: "juillet",
    8: "aout",
    9: "septembre",
    10: "octobre",
    11: "novembre",
    12: "decembre",
}

_ISO_DATE_RE = re.compile(r"^\d{4}([-\/])\d{2}\1\d{2}$")


def _to_date(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        match = _ISO_DATE_RE.match(text)
        if match:
            sep = match.group(1)
            fmt = "%Y-%m-%d" if sep == "-" else "%Y/%m/%d"
            parsed = pd.to_datetime(text, format=fmt, errors="coerce")
        else:
            parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    else:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _format_time(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, dt.datetime):
        hour = value.hour
        minute = value.minute
    elif isinstance(value, dt.time):
        hour = value.hour
        minute = value.minute
    else:
        text = str(value).strip()
        if not text:
            return ""
        compact = text.replace(" ", "")
        if "h" in compact:
            return compact
        parts = compact.split(":")
        if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
            hour = int(parts[0])
            minute = int(parts[1])
        elif compact.isdigit():
            hour = int(compact)
            minute = 0
        else:
            return compact
    return f"{hour}h" if minute == 0 else f"{hour}h{minute:02d}"


def _format_full_date(date_obj: dt.date | None) -> str:
    if not date_obj:
        return ""
    weekday = WEEKDAY_FR.get(date_obj.weekday(), "")
    month = date_obj.month
    if not weekday or not month:
        return ""
    return f"{weekday[0:3]} {date_obj.day}/{month:0>2}"


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _is_vo(value: object) -> bool:
    return _normalize_text(value).startswith("vo")


def _is_scolaire(categorie: object) -> bool:
    return bool(re.search(r"\bscol(?:aire)?\b", _normalize_text(categorie)))

def _is_jp(categorie: object) -> bool:
    return bool(re.search(r"\bjeune public\b", _normalize_text(categorie)))
def _is_Doc(categorie: object) -> bool:
    return bool(re.search(r"\bdoc\b", _normalize_text(categorie)))
def _is_Cgout(categorie: object) -> bool:
    return bool(re.search(r"\bcine gouter\b", _normalize_text(categorie)))
def _is_Cdisc(categorie: object) -> bool:
    return bool(re.search(r"\bcine discussion\b", _normalize_text(categorie)))
def _is_Cjeun(categorie: object) -> bool:
    return bool(re.search(r"\bcine jeunes\b", _normalize_text(categorie)))
def _is_CDoc(categorie: object) -> bool:
    return bool(re.search(r"\bcine coc\b", _normalize_text(categorie)))
def _is_CPat(categorie: object) -> bool:
    return bool(re.search(r"\bcine patrimoine\b", _normalize_text(categorie)))

def _format_categorie(categorie: object, vo) :
    label=""
    if _is_scolaire(categorie) :
        label='SCOL'
    elif _is_jp(categorie) :
        label='JP'
    elif _is_Doc(categorie) :
        label='Doc'
    if _is_Cgout(categorie) :
        label='CGout'
    elif _is_Cdisc(categorie) :
        label='Cdisc'
    elif _is_Cjeun(categorie) :
        label='Cjeun'
    elif _is_CPat(categorie) :
        label='CPat'
    elif _is_CDoc(categorie) :
        label='CDoc'

    return f"{label} {vo}".strip()

def main() -> int:
    df = pd.read_excel(IN_PATH, sheet_name=0, dtype=object).fillna("")
    rows = []
    for _, row in df.iterrows():
        date_value = row.get("Date", "")
        date_obj = _to_date(date_value)
        date_label = _format_full_date(date_obj)
        heure = row.get("Heure", "")

        titre = str(row.get("Titre", "")).strip()
        vovf = str(row.get("VOVF", "")).strip()
        if not vovf:
            vovf = str(row.get("Version", "")).strip()
        vo_label = "VO" if _is_vo(vovf) else ""
        categorie= _format_categorie(str(row.get("Categorie", "")).strip(),vo_label)
        rows.append([ date_label, heure, titre, categorie])

    out_df = pd.DataFrame(rows, columns=["Date","Heure", "Titre", "Categorie"])

    try:
        out_df.to_excel(OUT_PATH, index=False)

        wb = load_workbook(OUT_PATH)
        ws = wb.active
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        base_font = Font(name="Colibri", size=14, bold=False)

        for row in ws.iter_rows():
            for cell in row:
                cell.font = base_font
                if cell.column <= 2:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 56
        ws.column_dimensions["D"].width = 12

        wb.save(OUT_PATH)
    except PermissionError as exc:
        raise SystemExit(
            f"Impossible d'ecrire {OUT_PATH}. Ferme le fichier Excel s'il est ouvert, puis relance le script."
        ) from exc

    print(f"OK: {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
