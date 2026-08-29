from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


TOOLS_DIR = Path(__file__).resolve().parents[1]
if str(TOOLS_DIR) not in sys.path:
    sys.path.insert(0, str(TOOLS_DIR))

from google_sheet_source import (  # noqa: E402
    SourcePreparationError,
    prepare_source_workbook,
    select_programme_sheet,
)


def add_valid_programme_sheet(workbook: Workbook, title: str):
    worksheet = workbook.create_sheet(title)
    worksheet["E3"] = "URL allocine"
    worksheet["F3"] = "Titre"
    worksheet["A4"] = "Mercredi"
    worksheet["B4"] = "2-sept."
    worksheet["C5"] = "21h"
    worksheet["E5"] = "https://www.allocine.fr/film/fichefilm_gen_cfilm=1.html"
    worksheet["F5"] = f"Film {title}"
    worksheet["F5"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
    worksheet["E8"] = "Prochainement"
    worksheet["E9"] = "Film A, Film B"
    return worksheet


class ProgrammeSheetSelectionTests(unittest.TestCase):
    def test_latest_numeric_sheet_ignores_template(self):
        self.assertEqual(
            select_programme_sheet(["Template", "358", "359"]),
            (359, "359"),
        )

    def test_explicit_programme_is_selected(self):
        self.assertEqual(
            select_programme_sheet(["Template", "358", "359"], 358),
            (358, "358"),
        )

    def test_missing_explicit_programme_lists_available_tabs(self):
        with self.assertRaisesRegex(SourcePreparationError, "358, 359"):
            select_programme_sheet(["Template", "358", "359"], 360)


class SourceWorkbookPreparationTests(unittest.TestCase):
    def test_latest_programme_is_saved_as_feuil1_and_keeps_prochainement(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source = temp_path / "download.xlsx"
            destination = temp_path / "source.xlsx"

            workbook = Workbook()
            workbook.remove(workbook.active)
            add_valid_programme_sheet(workbook, "358")
            add_valid_programme_sheet(workbook, "359")
            workbook.create_sheet("Template")
            workbook.save(source)

            prepared = prepare_source_workbook(source, destination)

            self.assertEqual(prepared.programme, 359)
            self.assertEqual(prepared.original_sheet_name, "359")
            output = load_workbook(destination, data_only=True)
            self.assertEqual(output.sheetnames, ["Feuil1"])
            self.assertEqual(output["Feuil1"]["F5"].value, "Film 359")
            self.assertEqual(output["Feuil1"]["F5"].fill.fgColor.rgb, "00FF0000")
            self.assertEqual(output["Feuil1"]["E8"].value, "Prochainement")
            self.assertEqual(output["Feuil1"]["E9"].value, "Film A, Film B")

    def test_legacy_feuil1_is_accepted_for_local_sources(self):
        self.assertEqual(
            select_programme_sheet(
                ["Feuil1"],
                allow_legacy_sheet=True,
            ),
            (None, "Feuil1"),
        )

    def test_invalid_headers_are_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source = temp_path / "invalid.xlsx"
            destination = temp_path / "source.xlsx"
            workbook = Workbook()
            workbook.active.title = "359"
            workbook.save(source)

            with self.assertRaisesRegex(SourcePreparationError, "E3"):
                prepare_source_workbook(source, destination)


if __name__ == "__main__":
    unittest.main()
