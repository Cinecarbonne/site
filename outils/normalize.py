#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Normalize Ciné Carbonne Excel (Feuil1) into a clean table (v3: direct column mapping).

Entrée :  input/source.xlsx  (Feuil1)
Sorties :
    - work/normalized.xlsx
    - work/prochainement.json  (liste de textes "prochainement")
"""
from contextlib import nullcontext
from pathlib import Path
import re
from datetime import datetime, time
from dateutil import parser as dtparser
import json
import unicodedata

import pandas as pd
from openpyxl.styles import numbers
from openpyxl import load_workbook
from openpyxl.styles import Border, Side,Font, Alignment


# --- chemins ---
BASE_DIR            = Path(__file__).resolve().parent
INPUT_PATH          = BASE_DIR / "input/source.xlsx"
OUTPUT_PATH         = BASE_DIR / "work/normalized.xlsx"
PROCHAINEMENT_PATH  = BASE_DIR / "work/prochainement.json"
SHEET_NAME          = "Feuil1"

# --- colonnes du fichier source (index 0-based pour pandas) ---
COL_A, COL_B, COL_C = 0, 1, 2
COL_URL          = 4   # E
COL_TITRE        = 5   # F
COL_VERSION      = 6   # G
COL_CM           = 7   # H
COL_REAL         = 8   # I
COL_PRIX_INVITES = 9   # J
COL_CATEG        = 10  # K
COL_TARIF        = 11  # L
COL_COMMENT      = 12  # M

WEEKDAYS_FR = {"lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"}

# --- contexte date d'exécution ---
EXEC_DATE   = datetime.today().date()
IS_DECEMBER = (EXEC_DATE.month == 12)


# ------------------------------------------------------------
# utilitaires
# ------------------------------------------------------------

def is_weekday_label(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return False
    s = str(x).strip().lower()
    return any(s.startswith(w) for w in WEEKDAYS_FR)


def parse_date_cell(x):
    """Parse une cellule de date Excel (ou texte) en date, avec correction décembre → janvier N+1."""
    if pd.isna(x):
        return None

    # 1) Cas : vrai datetime / Timestamp venant d'Excel
    if isinstance(x, (pd.Timestamp, datetime)):
        d = pd.to_datetime(x).date()

        # Ajustement "décembre → janvier N+1"
        if IS_DECEMBER and d.month == 1 and d.year == EXEC_DATE.year:
            d = d.replace(year=EXEC_DATE.year + 1)

        return d

    # 2) Cas : texte à parser
    s = str(x).strip()
    for dayfirst in (True, False):
        try:
            d = dtparser.parse(s, dayfirst=dayfirst, fuzzy=True).date()

            # Même ajustement pour le texte
            if IS_DECEMBER and d.month == 1 and d.year == EXEC_DATE.year:
                d = d.replace(year=EXEC_DATE.year + 1)

            return d
        except Exception:
            pass

    return None

def parse_time_cell(x):
    if pd.isna(x):
        return None
    if isinstance(x, (pd.Timestamp, datetime)):
        t = pd.to_datetime(x).time()
        return t.replace(second=0, microsecond=0)
    if isinstance(x, time):
        return x.replace(second=0, microsecond=0)

    s = str(x).strip()
    m = re.search(r"(\d{1,2})\s*[h:]\s*(\d{1,2})?$", s)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2)) if m.group(2) else 0
        if 0 <= hh < 24 and 0 <= mm < 60:
            return time(hh, mm)

    try:
        t = dtparser.parse(s).time()
        return t.replace(second=0, microsecond=0)
    except Exception:
        return None


def norm_str(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    s = str(x).strip()
    return s or None


def extract_prochainement_blocks(raw):
    """Return upcoming text blocks from explicit markers or a trailing title list."""
    upcoming_blocks = []
    index_list = list(raw.index)
    trailing_candidates = []

    for pos, idx in enumerate(index_list):
        row = raw.loc[idx]
        a = row.get(COL_A)
        b = row.get(COL_B)
        c = row.get(COL_C)

        titre_cell_value = row.get(COL_URL)
        titre_norm = norm_str(titre_cell_value)

        has_weekday = is_weekday_label(a)
        has_date = parse_date_cell(b) is not None
        has_time = parse_time_cell(c) is not None

        if not (titre_norm and "prochainement" in titre_norm.lower() and not has_weekday and not has_date):
            if (
                titre_norm
                and not has_weekday
                and not has_date
                and not has_time
                and ("," in titre_norm or ";" in titre_norm or "\n" in titre_norm)
            ):
                trailing_candidates.append(titre_norm)
            continue

        next_pos = pos + 1
        if next_pos >= len(index_list):
            continue
        next_idx = index_list[next_pos]
        next_row = raw.loc[next_idx]
        next_title = norm_str(next_row.get(COL_TITRE))
        if next_title:
            upcoming_blocks.append(next_title)

    if not upcoming_blocks and trailing_candidates:
        upcoming_blocks.append(trailing_candidates[-1])

    return upcoming_blocks


def normalize_version(v):
    v = (v or "").strip().upper()
    if v == "VOSTFR":
        return "VOstFR"
    if v in ("VO", "VF"):
        return v
    return "VF"


def clean_label(value):
    text = str(value or "")
    text = text.replace("\u00A0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,;-")


def normalize_text_key(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def merge_comment_parts(*parts):
    merged = []
    seen = set()
    for part in parts:
        text = clean_label(part)
        if not text:
            continue
        key = normalize_text_key(text)
        if key and key not in seen:
            seen.add(key)
            merged.append(text)
    return " / ".join(merged)


def is_rewards_text(value):
    text = normalize_text_key(value)
    if not text:
        return False
    keywords = [
        "prix",
        "cesar",
        "oscar",
        "golden globe",
        "golden globes",
        "festival",
        "palme",
        "ours",
        "gan",
        "award",
        "laureat",
        "meilleur",
        "meilleure",
        "jury",
        "selection officielle",
        "mostra",
        "berlinale",
        "venise",
    ]
    return any(keyword in text for keyword in keywords)


def split_labels(value):
    text = clean_label(value)
    if not text:
        return []
    parts = [p.strip() for p in re.split(r"[;,]+", text) if p and p.strip()]
    return parts if parts else [text]


def format_money_token(value):
    raw = str(value or "").replace(",", ".")
    try:
        amount = float(raw)
    except Exception:
        return str(value or "")
    if amount.is_integer():
        return f"{int(amount)} \u20ac"
    return f"{amount:.2f}".replace(".", ",") + " \u20ac"


def normalize_tarif_field(tarif):
    text = clean_label(tarif)
    if not text:
        return "", False

    jp_in_tarif = bool(re.search(r"\bJP\b", text, flags=re.IGNORECASE))
    text = re.sub(r"\bJP\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bTU\b", "Tarif Unique", text, flags=re.IGNORECASE)
    text = re.sub(r"\bADH\b|\bADH(?=\d)", "Adh\u00e9rent", text, flags=re.IGNORECASE)

    text = re.sub(r"([A-Za-z])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([A-Za-z])", r"\1 \2", text)
    text = re.sub(r"\s*-\s*", " - ", text)
    text = re.sub(r"\s*/\s*", " / ", text)
    text = re.sub(
        r"\b(\d+(?:[.,]\d+)?)\b(?!\s*(?:ans?\b|h\b|heure?s?\b)|/\d)\s*(?:€|euros?)?",
        lambda m: format_money_token(m.group(1)),
        text,
        flags=re.IGNORECASE,
    )
    text = clean_label(text)

    return text, jp_in_tarif


def normalize_categorie_field(categorie, jp_in_tarif=False):
    labels = []
    seen = set()
    jp_found = bool(jp_in_tarif)

    for part in split_labels(categorie):
        token = clean_label(part)
        if not token:
            continue

        if re.search(r"\bJP\b", token, flags=re.IGNORECASE):
            jp_found = True
            token = re.sub(r"\bJP\b", "", token, flags=re.IGNORECASE)

        token = re.sub(r"\bDOC\b", "Documentaire", token, flags=re.IGNORECASE)
        if re.search(r"\bSCOL\b", token, flags=re.IGNORECASE):
            token = "Scolaire"

        token = clean_label(token)
        if not token:
            continue

        key = token.lower()
        if key not in seen:
            seen.add(key)
            labels.append(token)

    if jp_found and "jeune public" not in seen:
        labels.append("Jeune Public")

    return ", ".join(labels)


TITLE_MARKER_RE = re.compile(
    r"(^|[\s\-/,(])(?P<marker>VOSTFR|VOSTF|VOstFR|VO|VF|JP|SCOL|SCOLAIRE)(?=$|[\s\-/),])",
    flags=re.IGNORECASE,
)


def append_category_label(categorie, label):
    labels = split_labels(categorie)
    key = normalize_text_key(label)
    if key and all(normalize_text_key(existing) != key for existing in labels):
        labels.append(label)
    return ", ".join(labels)


def normalize_title_field(title):
    text = clean_label(title)
    version = None
    jp_in_title = False
    scol_in_title = False

    def replace_marker(match):
        nonlocal version, jp_in_title, scol_in_title
        marker = match.group("marker").upper()
        prefix = match.group(1) or ""
        if marker in {"VO", "VOSTF", "VOSTFR"}:
            version = "VOstFR" if marker != "VO" else "VO"
        elif marker == "VF":
            version = "VF"
        elif marker == "JP":
            jp_in_title = True
        elif marker in {"SCOL", "SCOLAIRE"}:
            scol_in_title = True
        return prefix

    cleaned = TITLE_MARKER_RE.sub(replace_marker, text)
    cleaned = clean_label(cleaned)
    return cleaned, version, jp_in_title, scol_in_title


CM_REF_RE = re.compile(r"\bCM\s*(\d+)\b", flags=re.IGNORECASE)
CM_DURATION_RE = re.compile(r"(\d+\s*(?:'|min)\s*\d{0,2})", flags=re.IGNORECASE)


def parse_cm_refs(value):
    refs = []
    seen = set()
    for number in CM_REF_RE.findall(str(value or "")):
        ref = f"CM{number}"
        if ref not in seen:
            seen.add(ref)
            refs.append(ref)
    return refs


def parse_cm_definition(value):
    text = clean_label(value)
    if not text:
        return None

    match = re.match(r"^\s*(CM\s*\d+)\s*:\s*(.+?)\s*$", text, flags=re.IGNORECASE)
    if not match:
        return None

    code = match.group(1).upper().replace(" ", "")
    rest = clean_label(match.group(2))
    parts = [clean_label(part) for part in re.split(r"\s+-\s+", rest) if clean_label(part)]

    titre = parts[0] if parts else rest
    genre = parts[1] if len(parts) > 1 else ""

    duration = ""
    if len(parts) > 2:
        duration_source = " - ".join(parts[2:])
        duration_match = CM_DURATION_RE.search(duration_source)
        if duration_match:
            duration = clean_label(duration_match.group(1))
    if not duration and len(parts) > 1:
        duration_match = CM_DURATION_RE.search(parts[-1])
        if duration_match:
            duration = clean_label(duration_match.group(1))
            if len(parts) == 2 and genre == parts[-1]:
                genre = ""

    display_parts = [titre, genre, duration]
    return {
        "code": code,
        "titre": titre,
        "genre": genre,
        "duree": duration,
        "texte": " - ".join(part for part in display_parts if part),
    }


def extract_cm_catalog(raw):
    catalog = {}
    max_cols = raw.shape[1] if hasattr(raw, "shape") else 0
    for row_idx in (0, 1):
        if row_idx not in raw.index:
            continue
        row = raw.loc[row_idx]
        for col_idx in range(max_cols):
            parsed = parse_cm_definition(row.get(col_idx))
            if parsed:
                catalog[parsed["code"]] = parsed
    return catalog


def resolve_courts_metrages(catalog, cm_value):
    resolved = []
    for ref in parse_cm_refs(cm_value):
        payload = dict(catalog.get(ref) or {})
        if not payload:
            payload = {
                "code": ref,
                "titre": "",
                "genre": "",
                "duree": "",
                "texte": ref,
            }
        resolved.append(payload)
    return resolved


def is_red_background(cell):
    """Retourne True si la cellule Excel a un fond rouge (#FF0000)."""
    fill = cell.fill
    if fill is None:
        return False

    fg = fill.fgColor
    if fg is None:
        return False

    rgb = getattr(fg, "rgb", None)
    if rgb is None:
        return False

    # rgb peut être une string ou un objet avec .rgb
    if not isinstance(rgb, str):
        rgb = getattr(rgb, "rgb", None)
        if rgb is None or not isinstance(rgb, str):
            return False

    code = rgb.upper().lstrip("#")
    # ARGB -> RRGGBB
    if len(code) == 8:
        code = code[2:]

    return code == "FF0000"


# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------

def main():
    if not INPUT_PATH.exists():
        raise SystemExit(f"[ERREUR] Fichier introuvable : {INPUT_PATH}")

    # pandas : valeurs
    raw = pd.read_excel(INPUT_PATH, sheet_name=SHEET_NAME, header=None, dtype=object)

    cm_catalog = extract_cm_catalog(raw)

    # openpyxl : styles (couleurs)
    wb = load_workbook(INPUT_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    records = []
    upcoming_blocks = extract_prochainement_blocks(raw)
    current_date = None

    for idx, row in raw.iterrows():
        # --------------------------------------------------------
        # 1) Mise à jour du jour courant
        # --------------------------------------------------------
        a = row.get(COL_A)
        b = row.get(COL_B)
        row_date = parse_date_cell(b)
        if is_weekday_label(a) and row_date:
            current_date = row_date

        # --------------------------------------------------------
        # 2) Détection séance classique
        #    Règle: on avance d'abord par date, puis on ne s'intéresse
        #    qu'aux lignes qui portent un horaire en colonne C.
        # --------------------------------------------------------
        t = parse_time_cell(row.get(COL_C))
        if not current_date or not t:
            continue

        titre = norm_str(row.get(COL_TITRE))
        if not titre:
            continue
        titre, version_in_title, jp_in_title, scol_in_title = normalize_title_field(titre)
        if not titre:
            continue

        titre_cell_excel = ws.cell(row=idx + 1, column=COL_TITRE + 1)  # openpyxl = 1-based
        if is_red_background(titre_cell_excel):
            continue

        version = normalize_version(norm_str(row.get(COL_VERSION)))
        if version_in_title:
            version = version_in_title
        cm = norm_str(row.get(COL_CM))
        courts_metrages = resolve_courts_metrages(cm_catalog, cm)
        realisateur =  norm_str(row.get(COL_REAL))
        prix_invites = norm_str(row.get(COL_PRIX_INVITES))
        categorie = norm_str(row.get(COL_CATEG))
        tarif = norm_str(row.get(COL_TARIF))
        commentaire = norm_str(row.get(COL_COMMENT))
        url_allocine = norm_str(row.get(COL_URL))
        recompenses = None

        if prix_invites:
            if is_rewards_text(prix_invites):
                recompenses = prix_invites
            else:
                commentaire = merge_comment_parts(prix_invites, commentaire)

        # --- Normalisation categorie/tarif ---
        tarif, jp_in_tarif = normalize_tarif_field(tarif)
        categorie = normalize_categorie_field(categorie, jp_in_tarif=jp_in_tarif or jp_in_title)
        if scol_in_title:
            categorie = append_category_label(categorie, "Scolaire")

        records.append({
            # "Date": current_date.strftime("%Y-%m-%d"),
            "Date": current_date,
            "Heure": f"{t.hour:02d}:{t.minute:02d}",
            "Titre": titre,
            "Version": version,
            "CM": cm,
            "courts_metrages": json.dumps(courts_metrages, ensure_ascii=False) if courts_metrages else "",
            "Realisateur" : realisateur,
            "Recompenses": recompenses,
            "Categorie": categorie,
            "Tarif": tarif,
            "Commentaire": commentaire,
            "url_allocine": url_allocine,
        })

    # --------------------------------------------------------
    # export des séances
    # --------------------------------------------------------
    df = pd.DataFrame(records, columns=[
        "Date", "Heure", "Titre", "Version", "CM", "courts_metrages", 'Realisateur',
        "Recompenses", "Categorie", "Tarif", "url_allocine", ""
    ])

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_PATH, index=False)
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    for c in ws['A:A'] :
        c.number_format = "DD/MM/YYYY"
    #ws.column_dimensions["A"].number_format = "DD/MM/YYYY"
    ws.column_dimensions["A"].auto_size = True
    ws.column_dimensions["B"].auto_size = True
    ws.column_dimensions["C"].width = 50

    ws.column_dimensions["H"].witdth = 30
    ws.column_dimensions["I"].witdth = 30

    # ajout d'une bordure Bleu autour de 3 serie de film pour les  caissiere

    def set_border(ws, cell_range):
        Large = Side(border_style="double", color="000000FF")
        rows = ws[cell_range]
        for row in rows:
            row[0].border  = Border(left=Large)
            row[-1].border = Border(right=Large)
        for c in rows[0]:
             c.border = Border(top=Large)
        for c in rows[-1]:
            c.border = Border(bottom=Large)

    nb_row_per_series=int ((ws.max_row-1)/3)
    ws.column_dimensions["L"].witdth = 40
    ws.column_dimensions["L"].alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions["L"].font  = Font(name='Calibri',
                size=14,
                bold=True,
                color='000000FF')
    for index in [0,1,2] :
        label=f'Saisie {index+1}'
        first_row=1++index+index*nb_row_per_series
        last_row=first_row+nb_row_per_series
        try :
            ws.unmerge_cells(f"L{first_row}:L{last_row}")
        except:
            print('')
        ws.cell(first_row,12).value=label
        ws.merge_cells(f"L{first_row}:L{last_row}")
        set_border(ws, f'A{first_row}:L{last_row}')

    wb.save( OUTPUT_PATH)

    print(f"[done] Ecrit : {OUTPUT_PATH} ({len(df)} lignes)")

    # --------------------------------------------------------
    # export "prochainement"
    # on écrit systématiquement un JSON (liste de chaînes)
    # --------------------------------------------------------
    PROCHAINEMENT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with PROCHAINEMENT_PATH.open("w", encoding="utf-8") as f:
        json.dump(upcoming_blocks, f, ensure_ascii=False, indent=2)

    print(f"[done] Ecrit : {PROCHAINEMENT_PATH} ({len(upcoming_blocks)} bloc(s))")


if __name__ == "__main__":
    main()
