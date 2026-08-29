#!/usr/bin/env python3
from __future__ import annotations

import re
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook


DEFAULT_SPREADSHEET_ID = "16Wj7AykT6a8fPc9Rtxmci4ondVT5COhw8L9aizZGMiU"
DEFAULT_SPREADSHEET_URL = (
    f"https://docs.google.com/spreadsheets/d/{DEFAULT_SPREADSHEET_ID}/edit"
)
NORMALIZED_SHEET_NAME = "Feuil1"
DOWNLOAD_TIMEOUT_SECONDS = 90


class SourcePreparationError(RuntimeError):
    """Raised when the monthly source workbook cannot be prepared safely."""


@dataclass(frozen=True)
class PreparedSource:
    path: Path
    programme: int | None
    original_sheet_name: str


def _normalized_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _programme_candidates(sheet_names: list[str]) -> list[tuple[int, str]]:
    candidates: list[tuple[int, str]] = []
    for sheet_name in sheet_names:
        stripped = sheet_name.strip()
        if re.fullmatch(r"\d+", stripped):
            candidates.append((int(stripped), sheet_name))
    return candidates


def select_programme_sheet(
    sheet_names: list[str],
    programme: int | None = None,
    *,
    allow_legacy_sheet: bool = False,
) -> tuple[int | None, str]:
    """Select an explicit programme tab or the greatest numeric tab."""
    candidates = _programme_candidates(sheet_names)

    if programme is not None:
        for number, sheet_name in candidates:
            if number == programme:
                return number, sheet_name
        available = ", ".join(str(number) for number, _ in sorted(candidates)) or "aucun"
        raise SourcePreparationError(
            f"Le programme {programme} n'existe pas dans le Google Sheet "
            f"(onglets numeriques disponibles : {available})."
        )

    if candidates:
        return max(candidates, key=lambda item: item[0])

    if allow_legacy_sheet and NORMALIZED_SHEET_NAME in sheet_names:
        return None, NORMALIZED_SHEET_NAME

    raise SourcePreparationError(
        "Aucun onglet de programme numerique n'a ete trouve dans le classeur."
    )


def validate_programme_sheet(worksheet) -> None:
    """Check the stable sentinels required by normalize.py."""
    url_header = _normalized_text(worksheet.cell(row=3, column=5).value)
    title_header = _normalized_text(worksheet.cell(row=3, column=6).value)
    if "url" not in url_header or "allocine" not in url_header:
        raise SourcePreparationError(
            "La cellule E3 doit contenir l'en-tete URL Allocine."
        )
    if title_header != "titre":
        raise SourcePreparationError("La cellule F3 doit contenir l'en-tete Titre.")

    session_count = 0
    for row in worksheet.iter_rows(min_row=4, min_col=3, max_col=6, values_only=True):
        hour, _, _, title = row
        if hour not in (None, "") and title not in (None, ""):
            session_count += 1
    if session_count == 0:
        raise SourcePreparationError(
            "Aucune seance avec horaire et titre n'a ete trouvee dans l'onglet."
        )


def _spreadsheet_id(spreadsheet_url_or_id: str) -> str:
    value = str(spreadsheet_url_or_id or "").strip()
    if re.fullmatch(r"[A-Za-z0-9_-]+", value):
        return value

    parsed = urlparse(value)
    match = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", parsed.path)
    if match:
        return match.group(1)
    raise SourcePreparationError(
        "Adresse Google Sheet invalide : impossible d'en extraire l'identifiant."
    )


def download_google_sheet(
    spreadsheet_url_or_id: str,
    destination: Path,
    *,
    timeout: int = DOWNLOAD_TIMEOUT_SECONDS,
) -> Path:
    spreadsheet_id = _spreadsheet_id(spreadsheet_url_or_id)
    export_url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    )
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)

    try:
        response = requests.get(export_url, timeout=timeout)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise SourcePreparationError(
            f"Impossible de telecharger le Google Sheet : {exc}"
        ) from exc

    if not response.content.startswith(b"PK"):
        raise SourcePreparationError(
            "Google n'a pas renvoye un fichier Excel. Verifier que le classeur "
            "est accessible en lecture."
        )
    destination.write_bytes(response.content)
    return destination


def prepare_source_workbook(
    workbook_path: Path,
    destination: Path,
    programme: int | None = None,
    *,
    allow_legacy_sheet: bool = False,
) -> PreparedSource:
    workbook_path = Path(workbook_path)
    destination = Path(destination)
    try:
        # Le Google Sheet utilise des formules pour plusieurs dates. On prepare
        # une photographie de publication a partir des valeurs calculees par
        # Google ; sauvegarder les formules avec openpyxl ferait perdre leur
        # cache et laisserait ces dates vides dans normalize.py.
        workbook = load_workbook(workbook_path, data_only=True)
    except Exception as exc:
        raise SourcePreparationError(
            f"Impossible d'ouvrir le classeur source {workbook_path}: {exc}"
        ) from exc

    selected_programme, selected_sheet_name = select_programme_sheet(
        workbook.sheetnames,
        programme,
        allow_legacy_sheet=allow_legacy_sheet,
    )
    selected_sheet = workbook[selected_sheet_name]
    validate_programme_sheet(selected_sheet)

    for worksheet in list(workbook.worksheets):
        if worksheet is not selected_sheet:
            workbook.remove(worksheet)
    selected_sheet.title = NORMALIZED_SHEET_NAME

    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix="cinecarbonne-source-",
        suffix=".xlsx",
        dir=destination.parent,
        delete=False,
    ) as temporary_file:
        temporary_path = Path(temporary_file.name)
    try:
        workbook.save(temporary_path)
        workbook.close()
        temporary_path.replace(destination)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()

    return PreparedSource(
        path=destination,
        programme=selected_programme,
        original_sheet_name=selected_sheet_name,
    )


def download_and_prepare_source(
    spreadsheet_url_or_id: str,
    destination: Path,
    programme: int | None = None,
) -> PreparedSource:
    with tempfile.TemporaryDirectory(prefix="cinecarbonne-google-sheet-") as temp_dir:
        downloaded_path = Path(temp_dir) / "programmes.xlsx"
        download_google_sheet(spreadsheet_url_or_id, downloaded_path)
        return prepare_source_workbook(downloaded_path, destination, programme)
